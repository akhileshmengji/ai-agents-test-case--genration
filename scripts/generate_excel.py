"""
generate_excel.py  -  QA workflow Excel writer

Reads a story JSON and writes a formatted .xlsx workbook with four sheets:
    1. "Manual Test Cases"   - 21-column enterprise test case sheet
    2. "Traceability Matrix" - auto-generated RTM
    3. "Requirement Gaps"    - gap analysis with severity colours
    4. "QA Review"           - smoke, regression, security, a11y, performance summary

Usage:
    python generate_excel.py <input.json> [output.xlsx]

If output.xlsx is omitted, saves to <output_dir>/<story_id>_test_cases.xlsx,
where <output_dir> comes from config.json (see config.example.json).
"""
import json
import os
import shutil
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qa_config import load_config

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="2F5496")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
SECTION_FONT = Font(bold=True, color="1F3864", size=11)
WRAP_TOP     = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="D9D9D9")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TYPE_FILL = {
    "Functional":      "E2EFDA",
    "Negative":        "FCE4D6",
    "Edge-Boundary":   "FFF2CC",
    "Non-Functional":  "DDEBF7",
    "E2E-Integration": "EDEDED",
    "Smoke":           "E8D5F5",
    "Regression":      "D5E8F5",
    "Security":        "F5D5D5",
    "Accessibility":   "D5F5E8",
    "Performance":     "F5EDD5",
    "Exploratory":     "F0F0F0",
}
SEVERITY_FILL = {
    "Critical": "C00000",
    "High":     "F4B0B0",
    "Medium":   "FCE4A6",
    "Low":      "D6E4BC",
}
PRIORITY_FILL = {
    "High":   "F4B0B0",
    "Medium": "FCE4A6",
    "Low":    "D6E4BC",
}
COVERAGE_FILL = {
    "Covered":           "D6E4BC",
    "Partially Covered": "FCE4A6",
    "Not Covered":       "F4B0B0",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, row, col, value="", fill=None, font=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BORDER
    cell.alignment = alignment or WRAP_TOP
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell


def _pfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


# ---------------------------------------------------------------------------
# Sheet 1 - Manual Test Cases (21 columns)
# ---------------------------------------------------------------------------
HEADERS_TC = [
    "S.No", "Module", "Sub Module", "Feature", "User Story ID",
    "Platform", "User Role", "Priority", "Severity",
    "Preconditions", "Test Case ID", "Test Scenario",
    "Test Case Description", "Test Data", "Test Steps",
    "Expected Result", "Actual Result", "Status", "Comments",
    "Requirement Mapping", "Automation Candidate",
]
COL_WIDTHS_TC = [6, 14, 14, 22, 14, 12, 14, 10, 10,
                 28, 12, 32, 32, 24, 44, 38, 18, 14, 14, 22, 18]


def build_test_cases_sheet(wb, data):
    ws = wb.active
    ws.title = "Manual Test Cases"
    ws.append(HEADERS_TC)
    _style_header(ws, len(HEADERS_TC))

    meta        = data.get("meta", {})
    story_id    = meta.get("story_id", "")
    def_module  = meta.get("module", "")
    def_feature = meta.get("feature", meta.get("title", ""))

    stats = defaultdict(int)

    for idx, tc in enumerate(data.get("test_cases", []), start=1):
        tc_type  = tc.get("type", "Functional")
        priority = tc.get("priority", "")
        severity = tc.get("severity") or priority  # fallback: severity = priority

        steps = tc.get("steps", "")
        if isinstance(steps, list):
            steps = "\n".join(f"{i}. {s}" for i, s in enumerate(steps, 1))

        row_vals = [
            idx,
            tc.get("module", def_module),
            tc.get("sub_module", ""),
            tc.get("feature", def_feature),
            story_id,
            tc.get("platform", "All"),
            tc.get("user_role", ""),
            priority,
            severity,
            tc.get("preconditions", ""),
            tc.get("id", ""),
            tc.get("test_scenario", tc.get("title", "")),
            tc.get("title", ""),
            tc.get("test_data", ""),
            steps,
            tc.get("expected_result", ""),
            "",              # Actual Result
            "Not Executed",  # Status
            "",              # Comments
            tc.get("requirement_mapping", tc.get("requirement_ref", "")),
            tc.get("automation_candidate", "TBD"),
        ]

        r = idx + 1
        for c, val in enumerate(row_vals, start=1):
            _cell(ws, r, c, val)

        # Priority colour (col 8)
        pfill = PRIORITY_FILL.get(priority)
        if pfill:
            ws.cell(row=r, column=8).fill = _pfill(pfill)
            ws.cell(row=r, column=8).alignment = CENTER

        # Severity colour (col 9)
        sfill = SEVERITY_FILL.get(severity)
        if sfill:
            ws.cell(row=r, column=9).fill = _pfill(sfill)
            ws.cell(row=r, column=9).alignment = CENTER
            if severity == "Critical":
                ws.cell(row=r, column=9).font = Font(bold=True, color="FFFFFF")

        # Test Scenario colour by type (col 12)
        tfill = TYPE_FILL.get(tc_type)
        if tfill:
            ws.cell(row=r, column=12).fill = _pfill(tfill)

        # Status column (col 18)
        ws.cell(row=r, column=18).alignment = CENTER
        ws.cell(row=r, column=18).fill = _pfill("EDEDED")

        stats[tc_type] += 1
        stats["total"] += 1

    _set_widths(ws, COL_WIDTHS_TC)
    return stats


# ---------------------------------------------------------------------------
# Sheet 2 - Requirement Traceability Matrix
# ---------------------------------------------------------------------------
def build_traceability_sheet(wb, data):
    ws = wb.create_sheet("Traceability Matrix")
    headers = ["AC / BR ID", "Requirement Description", "Mapped Test Case IDs", "Coverage Status"]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Build mapping: req_id -> [TC-IDs]
    mapping = defaultdict(list)
    for tc in data.get("test_cases", []):
        raw = tc.get("requirement_mapping", tc.get("requirement_ref", ""))
        if raw:
            for ref in [r.strip() for r in raw.replace(";", ",").split(",")]:
                if ref:
                    mapping[ref].append(tc.get("id", ""))

    # Collect refs from gaps too (for Not Covered detection)
    gap_lookup = {}
    for g in data.get("gaps", []):
        ref = g.get("req_ref", "").strip()
        if ref:
            gap_lookup[ref] = g.get("requirement_text", "")[:90]

    all_refs = sorted(set(list(mapping.keys()) + list(gap_lookup.keys())))

    for r_idx, ref in enumerate(all_refs, start=2):
        tcs = mapping.get(ref, [])
        tc_str = ", ".join(tcs) if tcs else "-"

        if len(tcs) == 0:
            status = "Not Covered"
        elif len(tcs) == 1:
            status = "Partially Covered"
        else:
            status = "Covered"

        desc = gap_lookup.get(ref, "")

        _cell(ws, r_idx, 1, ref, alignment=CENTER)
        _cell(ws, r_idx, 2, desc)
        _cell(ws, r_idx, 3, tc_str)
        cfill = COVERAGE_FILL.get(status)
        _cell(ws, r_idx, 4, status,
              fill=_pfill(cfill) if cfill else None,
              alignment=CENTER)

    _set_widths(ws, [18, 52, 38, 20])
    return len(all_refs)


# ---------------------------------------------------------------------------
# Sheet 3 - Requirement Gaps
# ---------------------------------------------------------------------------
def build_gaps_sheet(wb, data):
    ws = wb.create_sheet("Requirement Gaps")
    headers = ["Req Ref", "Requirement Text", "Gap Type", "Issue", "Suggested Fix", "Severity"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r_idx, g in enumerate(data.get("gaps", []), start=2):
        row_vals = [
            g.get("req_ref", ""), g.get("requirement_text", ""),
            g.get("gap_type", ""), g.get("issue", ""),
            g.get("suggested_fix", ""), g.get("severity", ""),
        ]
        for c, val in enumerate(row_vals, start=1):
            _cell(ws, r_idx, c, val)
        sfill = SEVERITY_FILL.get(g.get("severity", ""))
        if sfill:
            ws.cell(row=r_idx, column=6).fill = _pfill(sfill)
            ws.cell(row=r_idx, column=6).alignment = CENTER
            if g.get("severity") == "Critical":
                ws.cell(row=r_idx, column=6).font = Font(bold=True, color="FFFFFF")

    _set_widths(ws, [16, 42, 24, 44, 44, 12])
    return len(data.get("gaps", []))


# ---------------------------------------------------------------------------
# Sheet 4 - QA Review Summary
# ---------------------------------------------------------------------------
def build_qa_review_sheet(wb, data, stats):
    ws = wb.create_sheet("QA Review")
    meta     = data.get("meta", {})
    story_id = meta.get("story_id", "")
    title    = meta.get("title", "")
    quality  = meta.get("quality_score", "N/A")

    # ---- Auto-classify test cases into QA buckets --------------------------
    smoke_tcs        = []
    regression_tcs   = []
    security_tcs     = []
    a11y_tcs         = []
    perf_tcs         = []
    exploratory_tcs  = []

    for tc in data.get("test_cases", []):
        tc_type  = tc.get("type", "")
        tags     = [t.lower() for t in tc.get("tags", [])]
        title_l  = (tc.get("title", "") + " " + tc.get("test_scenario", "")).lower()
        priority = tc.get("priority", "")

        # Smoke: explicit type, tag, or high-priority functional
        if tc_type == "Smoke" or "smoke" in tags or (priority == "High" and tc_type == "Functional"):
            smoke_tcs.append(tc)

        # Regression: explicit type, tag, or E2E
        if tc_type in ("Regression", "E2E-Integration") or "regression" in tags:
            regression_tcs.append(tc)

        # Security
        if tc_type == "Security" or "security" in tags or any(
            w in title_l for w in ("security", "token", "session", "unauthori", "idor", "rate limit")
        ):
            security_tcs.append(tc)

        # Accessibility
        if tc_type == "Accessibility" or "accessibility" in tags or any(
            w in title_l for w in ("accessibility", "a11y", "screen reader", "keyboard", "focus", "aria", "wcag", "voiceover", "talkback")
        ):
            a11y_tcs.append(tc)

        # Performance / Non-Functional
        if tc_type in ("Performance", "Non-Functional") or "performance" in tags or any(
            w in title_l for w in ("performance", "load time", "p95", "latency", "speed")
        ):
            perf_tcs.append(tc)

        # Exploratory
        if tc_type == "Exploratory" or "exploratory" in tags:
            exploratory_tcs.append(tc)

    # ---- Sheet writing helpers ---------------------------------------------
    current_row = [1]  # mutable counter

    def _merge_header(label, cols=4):
        r = current_row[0]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        current_row[0] += 1

    def _sub_header(label, cols=4):
        r = current_row[0]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row[0] += 1

    def _row(*vals):
        r = current_row[0]
        for c, v in enumerate(vals, start=1):
            _cell(ws, r, c, v)
        current_row[0] += 1

    def _blank():
        current_row[0] += 1

    def _tc_rows(tc_list, col4="Platform"):
        _row("Test Case ID", "Test Scenario / Title", "Priority", col4)
        for tc in tc_list:
            _row(
                tc.get("id", ""),
                tc.get("test_scenario", tc.get("title", "")),
                tc.get("priority", ""),
                tc.get("platform", "All") if col4 == "Platform" else tc.get("type", ""),
            )
        _blank()

    # ---- Story summary -----------------------------------------------------
    _merge_header(f"QA Review Summary - {story_id}: {title}")
    _row("Quality Score", str(quality))
    _row("Total Test Cases", str(stats.get("total", 0)))
    _blank()

    # ---- Distribution table ------------------------------------------------
    _sub_header("Test Case Distribution")
    _row("Type", "Count")
    for tc_type in sorted(k for k in stats if k != "total"):
        _row(tc_type, str(stats[tc_type]))
    _blank()

    # ---- Negative / Boundary / Edge coverage check -------------------------
    neg_count   = stats.get("Negative", 0)
    edge_count  = stats.get("Edge-Boundary", 0)
    total       = stats.get("total", 1)
    nbe_pct     = round((neg_count + edge_count) / total * 100)
    coverage_ok = "Meets 30-40% threshold" if nbe_pct >= 30 else "Below 30% threshold - add more Negative/Boundary/Edge cases"
    _sub_header("Negative + Boundary + Edge Coverage")
    _row("Combined %", f"{nbe_pct}%", coverage_ok)
    _blank()

    # ---- Smoke tests -------------------------------------------------------
    _sub_header("Smoke Test Scenarios")
    if smoke_tcs:
        _tc_rows(smoke_tcs[:20])
    else:
        _row("No smoke test cases identified.")
        _blank()

    # ---- Regression tests --------------------------------------------------
    _sub_header("Regression Test Scenarios")
    if regression_tcs:
        _tc_rows(regression_tcs, col4="Type")
    else:
        _row("No regression / E2E test cases identified.")
        _blank()

    # ---- Security checks ---------------------------------------------------
    _sub_header("Security Checks")
    if security_tcs:
        _tc_rows(security_tcs, col4="Type")
    else:
        _row("No explicit security test cases found. Suggested checks:")
        _row("- Verify auth tokens cannot be replayed after use")
        _row("- Verify session expiry blocks mid-flow actions")
        _row("- Verify a user cannot access another user's protected resources (IDOR)")
        _row("- Verify rate limiting on sensitive endpoints (login, OTP, invite)")
        _blank()

    # ---- Accessibility checks ----------------------------------------------
    _sub_header("Accessibility Checks (WCAG 2.1 AA)")
    if a11y_tcs:
        _tc_rows(a11y_tcs, col4="Platform")
    else:
        _row("No explicit accessibility test cases found. Suggested checks:")
        _row("- All CTAs have accessible labels (VoiceOver / TalkBack)")
        _row("- Focus order is logical across all screens")
        _row("- Error messages announced assertively by screen reader")
        _row("- Keyboard-only navigation completes the full flow")
        _row("- Color contrast meets WCAG AA on all states")
        _blank()

    # ---- Performance checks ------------------------------------------------
    _sub_header("Performance Checks")
    if perf_tcs:
        _tc_rows(perf_tcs, col4="Type")
    else:
        _row("No explicit performance test cases found. Suggested checks:")
        _row("- Key screen loads within an acceptable budget on 4G")
        _row("- Primary API responds within an acceptable budget under normal load")
        _blank()

    # ---- Exploratory ideas -------------------------------------------------
    _sub_header("Exploratory Testing Ideas")
    if exploratory_tcs:
        for tc in exploratory_tcs:
            _row(tc.get("id", ""), tc.get("title", ""))
    else:
        _row("- Repeat the primary action multiple times rapidly - check for duplicate side effects")
        _row("- Open the flow in multiple browser tabs simultaneously")
        _row("- Test with extremely long input values in free-text fields")
        _row("- Simulate network drop between each step of the flow")
        _row("- Test with device language set to non-English")
        _row("- Test session/token expiry at every step of the flow")
    _blank()

    # ---- Extra qa_review section from JSON (if present) -------------------
    qa_review = data.get("qa_review", {})
    if qa_review:
        _sub_header("Additional QA Notes (from story)")
        for key, items in qa_review.items():
            _row(key.replace("_", " ").title())
            if isinstance(items, list):
                for item in items:
                    _row(f"  - {item}")
            _blank()

    _set_widths(ws, [24, 64, 18, 22])


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------
def build(data, out_path):
    wb = Workbook()
    stats  = build_test_cases_sheet(wb, data)
    n_rtm  = build_traceability_sheet(wb, data)
    n_gap  = build_gaps_sheet(wb, data)
    build_qa_review_sheet(wb, data, stats)
    wb.save(out_path)
    return stats, n_rtm, n_gap


def main():
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("Usage: python generate_excel.py <input.json> [output.xlsx]")
        sys.exit(1)

    in_path = sys.argv[1]
    with open(in_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    config = load_config()

    if len(sys.argv) == 3:
        out_path = sys.argv[2]
    else:
        story_id = data.get("meta", {}).get("story_id", "story")
        os.makedirs(config["output_dir"], exist_ok=True)
        out_path = os.path.join(config["output_dir"], f"{story_id}_test_cases.xlsx")

    stats, n_rtm, n_gap = build(data, out_path)
    n_tc = stats.get("total", 0)

    print(f"OK  ->  {out_path}")
    print(f"     {n_tc} test cases  |  {n_rtm} RTM entries  |  {n_gap} gaps")

    dist = {k: v for k, v in stats.items() if k != "total"}
    print(f"     Distribution: {dist}")

    # Optionally copy the source JSON to a downstream automation pipeline dir.
    # Disabled by default - set "automation_pipeline_dir" in config.json to enable.
    if config.get("automation_pipeline_dir"):
        os.makedirs(config["automation_pipeline_dir"], exist_ok=True)
        json_dest = os.path.join(config["automation_pipeline_dir"], os.path.basename(in_path))
        shutil.copy2(in_path, json_dest)
        print(f"JSON ->  {json_dest}")


if __name__ == "__main__":
    main()
